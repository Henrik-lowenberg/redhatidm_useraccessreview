#!/usr/bin/python
# Creator: Henrik Lowenberg
# Version: 0.5 Beta
#
# The purpose of this script
# is to extract all components
# that makes up a users access to a server
# that can serve as a user access review report
# this can be combined with AD groups and user accounts
# from MS AD
# Prerequisity:
# create local idm user api_reader
# create new role api-reader
# assign IPA Masters Readers to that role
# assign role to user
# login as user and set new pwd
# Read
# https://python-freeipa.readthedocs.io/en/latest/
#
# How-to install pre-requisite libraries:
# python -m pip install python-freeipa openpyxl
#
# Changelog:
# switch from csv to openpyxl library to create excel files directly
# rewrite to be object-oriented instead of procedural
# Better error handling
#
# ToDo:
# format data better, detect/convert lists to strings. 
#  json.loads to parse data?
#  before writing to Excel sheet.
# rewrite to use instance variables instead of global variables in class
# Comment more!
#
# Import the freeipa library
from python_freeipa import ClientMeta
# to work with Excel
from openpyxl import Workbook
# remove unwanted warnings
import urllib3


def main():
    # get rid of unverified HTTPS connection warnings
    urllib3.disable_warnings()
    # get login credentials as global vars
    client = ClientMeta("ipaserver.example.com", verify_ssl=False)
    client.login("api-user","somepwd")
        

    def listToString(s): # convert list to string
        # initialize an empty string
        str1 = ""
        # traverse in the string
        for ele in s:
            str1 += ele + ","
        
        return str1


    def getUsers(client): # Get all IDM users
        # use the freeipa module user_find method and store in userlist list
        userlist = client.user_find()
        # get sub category result
        userresult = userlist.get("result")
        # Create empty list
        listOfUsers = []
        # pad the empty list with a header line
        listOfUsers.append(['USER NAME', 'FULL NAME', 'GROUPS'])
        # loop through each line in list
        for row in userresult:
            # declare empty list where we inserts 3 elements of all user elements
            user = []
            user.append(row.get('uid')[0])
            user.append(row.get('cn')[0])
            user.append(listToString(row.get('memberof_group')))
            # add user list to bigger list of all users
            listOfUsers.append(user)
        
        return listOfUsers


    def getUserGroups(client): # Get all user groups
        #grouplist = client.group_find(posix=True)
        grouplist = client.group_find()
        groupresult = grouplist.get("result")
        # Create empty list
        listOfGroups = []
        listOfGroups.append(['GROUP NAME', 'GROUP MEMBERSHIP', 'SUDORULE', 'HBACRULE'])
        for row in groupresult:
            group = []
            group.append(row.get('cn')[0])
            group.append(row.get('memberof_group'))
        #    group.append(row.get('memberofindirect_sudorule'))
            group.append(row.get('memberof_sudorule'))
            group.append(row.get('memberof_hbacrule'))
            listOfGroups.append(group)
        
        return listOfGroups

        
    def getHostGroups(client): # Get all host groups
        hostgrouplist = client.hostgroup_find()
        hostgroupresult = hostgrouplist.get("result")
        # Create empty list
        listOfHostGroups = []
        listOfHostGroups.append(['HOSTGROUP NAME', 'HBACRULE MEMBERSHIP', 'HOSTS'])
        for row in hostgroupresult:
            hostgroup = []
            hostgroup.append(row.get('cn')[0])
            hostgroup.append(row.get('memberof_hbacrule'))
            hostgroup.append(row.get('member_host'))
            listOfHostGroups.append(hostgroup)
        
        return listOfHostGroups


    def getHbacRules(client): # get all HBAC rules
        hbacrulelist = client.hbacrule_find()
        hbacruleresult = hbacrulelist.get("result")
        # Create empty list
        listOfHbacRules = []
        listOfHbacRules.append(['HBACRULE NAME', 'MEMBER USER', 'TYPE OF ACCESS', 'MEMBER GROUP'])
        for row in hbacruleresult:
            hbacrule = []
            hbacrule.append(row.get('cn')[0])
            hbacrule.append(row.get('memberuser_user'))
            hbacrule.append(row.get('accessruletype'))
            hbacrule.append(row.get('memberuser_group'))
            listOfHbacRules.append(hbacrule)

        return listOfHbacRules


    def getSudoRules(client): # get all Sudo rules
        sudorulelist = client.sudorule_find()
        sudoruleresult = sudorulelist.get("result")
        # Create empty list
        listOfSudoRules = []
        listOfSudoRules.append(['SUDORULE NAME', 'RUN AS', 'MEMBER HOST', 'MEMBER USERGROUP', 'MEMBER HOSTGROUP', 'CMD CATEGORY'])
        for row in sudoruleresult:
            sudorule = []
            sudorule.append(row.get('cn')[0])
            sudorule.append(row.get('ipasudorunasextuser'))
            sudorule.append(row.get('memberhost'))
            sudorule.append(row.get('memberuser_group'))
            sudorule.append(row.get('memberhost_hostgroup'))
            sudorule.append(row.get('cmdcategory'))
            listOfSudoRules.append(sudorule)
        
        return listOfSudoRules


    def getSudoCmds(client): # get all Sudo commands
        sudocmdlist = client.sudocmd_find()
        sudocmdresult = sudocmdlist.get("result")
        # Create empty list
        listOfSudoCmds = []
        listOfSudoCmds.append(['SUDOCMD NAME', 'DESCRIPTION'])

        for row in sudocmdresult:
            sudocmd = []
            sudocmd.append(row.get('sudocmd')[0])
            sudocmd.append(row.get('description'))
            listOfSudoCmds.append(sudocmd)

        return listOfSudoCmds


    def getSudoCmdGroups(client): # get all Sudo command groups
        sudocmdgrouplist = client.sudocmdgroup_find()
        sudocmdresult = sudocmdgrouplist.get("result")
        # Create empty list
        listOfSudoCmdGroups = []
        listOfSudoCmdGroups.append(['SUDOCMDGROUP NAME', 'SUDO CMD'])
        for row in sudocmdresult:
            sudocmdgroup = []
            sudocmdgroup.append(row.get('cn'))
            sudocmdgroup.append(row.get('member_sudocmd'))
            listOfSudoCmdGroups.append(sudocmdgroup)

        return listOfSudoCmdGroups


    def dataToExcel(listOfUsers, listOfGroups, listOfHostGroups, listOfHbacRules, listOfSudoRules, listOfSudoCmds, listOfSudoCmdGroups):
        # create empty workbook
        wb = Workbook()
        # change name of worksheet
        ws1 = wb.active
        ws1.title = "users_sheet"
        print("users")
        # loop through list
        for row in listOfUsers:
            try: # check for nestled lists
                if(isinstance(row, str)): # add row if string
                    ws1.append(row)
                elif(isinstance(row, list)): # convert nestled lists to strings
                    row = listToString(row)
                    ws1.append(row)

            except Exception as e: 
                print(e)
                print(row)

        print("usergroup")
        ws2 = wb.create_sheet("usergroup_sheet", 0)
        ws2 = wb.active
        for row in listOfGroups:
            try:
                ws2.append(row)
            except Exception as e: 
                print(e)
                print(row)

        print("hostgroups")    
        ws3 = wb.create_sheet("hostgroups_sheet", 0)
        ws3 = wb.active
        for row in listOfHostGroups:
            try:
                ws3.append(row)
            except Exception as e: 
                print(e)
                print(row)

        print("HBACrules")
        ws4 = wb.create_sheet("hbacrules_sheet", 0)
        ws4 = wb.active
        for row in listOfHbacRules:
            try:
                ws4.append(row)
            except Exception as e: 
                print(e)
                print(row)

        print("sudorules")
        ws5 = wb.create_sheet("sudorules_sheet", 0)
        ws5 = wb.active
        for row in listOfSudoRules:
            try:
                ws5.append(row)
            except Exception as e: 
                print(e)
                print(row)

        print("sudocmds")
        ws6 = wb.create_sheet("sudocmds_sheet", 0)
        ws6 = wb.active
        for row in listOfSudoCmds:
            try:
                ws6.append(row)
            except Exception as e:
                print(e)
                print(row)

        print("sudocmdgroups")
        ws7 = wb.create_sheet("sudocmdgroups_sheet", 0)
        ws7 = wb.active
        for row in listOfSudoCmdGroups:
            try:
                ws7.append(row)
            except Exception as e: 
                print(e)
                print(row)

        print("saving file...")
        wb.save(filename = 'userAccessReview.xlsx')


    # Method Loader
    listOfUsers = getUsers(client)
    listOfGroups = getUserGroups(client)
    listOfHostGroups = getHostGroups(client)
    listOfHbacRules = getHbacRules(client)
    listOfSudoRules = getSudoRules(client)
    listOfSudoCmds = getSudoCmds(client)
    listOfSudoCmdGroups = getSudoCmdGroups(client)
    dataToExcel(listOfUsers, listOfGroups, listOfHostGroups, listOfHbacRules, listOfSudoRules, listOfSudoCmds, listOfSudoCmdGroups)

 
if __name__ == "__main__": 
# Load main class if script is
# executed directly (not sourced/included)
    main()


# END OF SCRIPT #
